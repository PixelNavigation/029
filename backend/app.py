import os
import json
import hashlib
import uuid
import re
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from supabase import create_client
from dotenv import load_dotenv
from datetime import datetime
from werkzeug.utils import secure_filename
from pathlib import Path
from ocr_pipeline import process_certificate_file, batch_process_and_save


load_dotenv()

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=True)

# File upload configuration
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'tiff', 'bmp', 'doc', 'docx', 'xls', 'xlsx', 'csv'}
MAX_CONTENT_LENGTH = 100 * 1024 * 1024  #100MB max file size

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

# Create uploads directory if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Create metadata file path
METADATA_FILE = os.path.join(UPLOAD_FOLDER, 'uploads_metadata.json')


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def save_metadata(metadata):
    """Save upload metadata to JSON file"""
    try:
        existing_data = []
        if os.path.exists(METADATA_FILE) and os.path.getsize(METADATA_FILE) > 0:
            try:
                with open(METADATA_FILE, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
            except json.JSONDecodeError:
                existing_data = []
        
        existing_data.append(metadata)
        
        with open(METADATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Error saving metadata: {str(e)}")


def get_all_metadata():
    """Retrieve all upload metadata"""
    try:
        if os.path.exists(METADATA_FILE) and os.path.getsize(METADATA_FILE) > 0:
            with open(METADATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return []
    except json.JSONDecodeError:
        print(f"Error reading metadata: Invalid JSON")
        return []
    except Exception as e:
        print(f"Error reading metadata: {str(e)}")
        return []


SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_ANON_KEY = os.environ.get("SUPABASE_ANON_KEY")
supabase = None

if SUPABASE_URL and SUPABASE_ANON_KEY:
    supabase = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)


@app.route("/health", methods=["GET"])
def health():
    return jsonify({
        "status": "ok",
        "supabase_configured": bool(supabase),
    })


@app.route("/api/auth/signup", methods=["POST"])
def signup():
    """Handle student signup - creates account in Supabase"""
    if not supabase:
        return jsonify({"error": "Database not configured"}), 503

    data = request.get_json() or {}
    
    # Validate required fields
    # Secure fields required to be stored in secure table
    required_fields = ['email', 'password', 'studentId', 'phone']
    missing_fields = [field for field in required_fields if not data.get(field)]
    
    if missing_fields:
        return jsonify({"error": f"Missing required fields: {', '.join(missing_fields)}"}), 400
    
    try:
        # Check if user already exists (by email) in secure table
        existing_user = supabase.table('students').select('*').eq('email', data['email']).execute()
        if existing_user.data and len(existing_user.data) > 0:
            return jsonify({"error": "User with this email already exists. Please sign in."}), 409

        # Check if student ID already exists in secure table
        existing_student_id = supabase.table('students').select('*').eq('student_id', data['studentId']).execute()
        if existing_student_id.data and len(existing_student_id.data) > 0:
            return jsonify({"error": "Student ID already exists"}), 409

        # Insert secure record into students table (only secure data)
        secure_student = {
            'email': data['email'],
            'password': data['password'],  # In production, hash this!
            'student_id': data['studentId'],
            'phone': data['phone'],
            'role': 'student',
            'verified': False,
            'created_at': datetime.utcnow().isoformat()
        }

        result_secure = supabase.table('students').insert(secure_student).execute()
        if not (result_secure.data and len(result_secure.data) > 0):
            return jsonify({"error": "Failed to create secure account record"}), 500

        # Insert public profile data into public_profiles table (non-sensitive)
        public_profile = {
            'student_id': data['studentId'],
            'name': data.get('name', ''),
            'university': data.get('university', ''),
            'course': data.get('course', ''),
            'year': data.get('year', ''),
            'profile_photo': data.get('profilePhoto', ''),
            'social_links': data.get('socialLinks', {}),
            'bio': data.get('bio', '')
        }

        result_public = supabase.table('public_profiles').insert(public_profile).execute()

        return jsonify({
            "success": True,
            "message": "Account created successfully! Please sign in with your credentials."
        }), 201
            
    except Exception as e:
        print(f"Signup error: {str(e)}")
        return jsonify({"error": f"Registration failed: {str(e)}"}), 500


@app.route("/api/upload-profile-photo", methods=["POST"])
def upload_profile_photo():
    """Upload profile photo to Supabase Storage"""
    if not supabase:
        return jsonify({"error": "Database not configured"}), 503

    if 'photo' not in request.files:
        return jsonify({"error": "No photo file provided"}), 400

    file = request.files['photo']
    student_id = request.form.get('studentId')

    if not student_id:
        return jsonify({"error": "Student ID is required"}), 400

    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    if file and allowed_file(file.filename):
        try:
            # Generate unique filename
            file_ext = file.filename.rsplit('.', 1)[1].lower()
            unique_filename = f"profile_photos/{student_id}_{uuid.uuid4().hex}.{file_ext}"
            
            # Read file content
            file_content = file.read()
            
            # Upload to Supabase Storage using correct Python API
            try:
                # The Python Supabase library expects bytes and path
                storage = supabase.storage.from_('student-profiles')
                
                # Upload file
                res = storage.upload(
                    path=unique_filename,
                    file=file_content,
                    file_options={"content-type": f"image/{file_ext}"}
                )
                
                print(f"Upload successful: {unique_filename}")
                
            except Exception as upload_err:
                print(f"Supabase upload error: {str(upload_err)}")
                print(f"Error type: {type(upload_err).__name__}")
                import traceback
                traceback.print_exc()
                
                # Try alternative upload method
                try:
                    res = storage.upload(path=unique_filename, file=file_content)
                    print("Upload succeeded with basic method")
                except Exception as retry_err:
                    print(f"Retry also failed: {str(retry_err)}")
                    raise
            
            # Get public URL
            public_url = supabase.storage.from_('student-profiles').get_public_url(unique_filename)
            print(f"Public URL: {public_url}")
            
            return jsonify({
                "success": True,
                "url": public_url
            }), 200

        except Exception as e:
            print(f"Upload error: {str(e)}")
            print(f"Error type: {type(e).__name__}")
            import traceback
            traceback.print_exc()
            return jsonify({"error": f"Failed to upload photo: {str(e)}"}), 500
    
    return jsonify({"error": "Invalid file type"}), 400


@app.route("/api/auth/signin", methods=["POST"])
def signin():
    """Handle student signin - authenticates user"""
    if not supabase:
        return jsonify({"error": "Database not configured"}), 503

    data = request.get_json() or {}
    email = data.get('email')
    password = data.get('password')
    role = data.get('role', 'student')

    if not email or not password:
        return jsonify({"error": "Email and password are required"}), 400

    try:
        # Fetch student from database
        result = supabase.table('students').select('*').eq('email', email).execute()
        if not result.data or len(result.data) == 0:
            return jsonify({"error": "Invalid email or password"}), 401

        student = result.data[0]
        # Verify password (in production, use proper password hashing comparison)
        if student.get('password') != password:
            return jsonify({"error": "Invalid email or password"}), 401

        # Fetch public profile for this student_id (if exists)
        public = supabase.table('public_profiles').select('*').eq('student_id', student.get('student_id')).maybe_single().execute()
        public_data = public.data if public and public.data else {}

        # Build response (exclude password)
        user_data = {
            'id': student.get('id'),
            'email': student.get('email'),
            'role': student.get('role', 'student'),
            'studentId': student.get('student_id'),
            'phone': student.get('phone'),
            'verified': student.get('verified', False),
            'createdAt': student.get('created_at'),
            'profile': {
                'name': public_data.get('name') if public_data else '',
                'university': public_data.get('university') if public_data else '',
                'course': public_data.get('course') if public_data else '',
                'year': public_data.get('year') if public_data else '',
                'profilePhoto': public_data.get('profile_photo') if public_data else f"https://ui-avatars.com/api/?name=User&background=3b82f6&color=fff",
                'socialLinks': public_data.get('social_links') if public_data else {}
            }
        }

        return jsonify({
            "success": True,
            "user": user_data
        }), 200
        
    except Exception as e:
        print(f"Signin error: {str(e)}")
        return jsonify({"error": f"Login failed: {str(e)}"}), 500


@app.route("/api/student/upload-certificate", methods=["POST"])
def upload_student_certificate():
    """Upload student certificate to Supabase Storage"""
    if not supabase:
        return jsonify({"error": "Database not configured"}), 503

    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files['file']
    student_id = request.form.get('studentId')
    document_type = request.form.get('documentType', 'certificate')

    if not student_id:
        return jsonify({"error": "Student ID is required"}), 400

    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    if file and allowed_file(file.filename):
        try:
            # Generate unique filename
            file_ext = file.filename.rsplit('.', 1)[1].lower()
            timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            unique_filename = f"certificates/{student_id}/{document_type}_{timestamp}_{uuid.uuid4().hex}.{file_ext}"
            
            # Read file content
            file_content = file.read()
            
            # Upload to Supabase Storage using correct Python API
            try:
                storage = supabase.storage.from_('student-profiles')
                
                content_type = f'application/{file_ext}' if file_ext == 'pdf' else f'image/{file_ext}'
                res = storage.upload(
                    path=unique_filename,
                    file=file_content,
                    file_options={"content-type": content_type}
                )
                
                print(f"Certificate upload successful: {unique_filename}")
                
            except Exception as upload_err:
                print(f"Supabase certificate upload error: {str(upload_err)}")
                import traceback
                traceback.print_exc()
                
                # Try without options
                try:
                    res = storage.upload(path=unique_filename, file=file_content)
                    print("Certificate upload succeeded with basic method")
                except Exception as retry_err:
                    print(f"Certificate retry failed: {str(retry_err)}")
                    raise
            
            # Get public URL
            public_url = supabase.storage.from_('student-profiles').get_public_url(unique_filename)
            
            # Store certificate metadata in database
            certificate_data = {
                'student_id': student_id,
                'document_type': document_type,
                'file_name': file.filename,
                'file_url': public_url,
                'file_size': len(file_content),
                'uploaded_at': datetime.utcnow().isoformat(),
                'verification_status': 'pending'
            }
            
            # Insert into certificates table
            try:
                cert_result = supabase.table('certificates').insert(certificate_data).execute()
                certificate_id = cert_result.data[0]['id'] if cert_result.data else None
            except Exception as db_err:
                print(f"Database insert error: {str(db_err)}")
                # Continue even if database insert fails - at least file is uploaded
                certificate_id = None
            
            return jsonify({
                "success": True,
                "url": public_url,
                "certificateId": certificate_id,
                "fileName": file.filename
            }), 200

        except Exception as e:
            print(f"Upload error: {str(e)}")
            return jsonify({"error": f"Failed to upload certificate: {str(e)}"}), 500
    
    return jsonify({"error": "Invalid file type"}), 400


@app.route("/api/student/certificates/<student_id>", methods=["GET"])
def get_student_certificates(student_id):
    """Get all certificates for a student"""
    if not supabase:
        return jsonify({"error": "Database not configured"}), 503

    try:
        result = supabase.table('certificates').select('*').eq('student_id', student_id).execute()
        
        return jsonify({
            "success": True,
            "certificates": result.data
        }), 200

    except Exception as e:
        print(f"Fetch error: {str(e)}")
        return jsonify({"error": f"Failed to fetch certificates: {str(e)}"}), 500


@app.route("/api/institution/upload-certificates", methods=["POST"])
def upload_certificates():
    """Handle bulk certificate uploads from institutions"""
    if 'files[]' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files[]')
    institution_id = request.form.get('institution_id', 'unknown')
    institution_name = request.form.get('institution_name', 'Unknown Institution')
    
    if not files:
        return jsonify({"error": "No files selected"}), 400

    uploaded_files = []
    failed_files = []
    
    # Sanitize institution name for folder/file naming
    safe_institution_name = re.sub(r'[^\w\s-]', '', institution_name)
    safe_institution_name = re.sub(r'[-\s]+', '_', safe_institution_name)
    
    # Create institution-specific subfolder in files directory
    files_folder = os.path.join(UPLOAD_FOLDER, 'files', safe_institution_name)
    os.makedirs(files_folder, exist_ok=True)
    
    # Get current date and time
    now = datetime.now()
    upload_date = now.strftime('%Y%m%d')
    upload_time = now.strftime('%H%M%S')
    batch_id = f"{upload_date}_{upload_time}"

    for file in files:
        if file and file.filename:
            if allowed_file(file.filename):
                try:
                    # Get original filename and extension
                    original_filename = secure_filename(file.filename)
                    name_without_ext, file_extension = os.path.splitext(original_filename)
                    
                    # Create new filename: InstitutionName_YYYYMMDD_HHMMSS_OriginalName.ext
                    new_filename = f"{safe_institution_name}_{upload_date}_{upload_time}_{name_without_ext}{file_extension}"
                    
                    # Save the file
                    file_path = os.path.join(files_folder, new_filename)
                    file.save(file_path)
                    
                    # Get file info
                    file_size = os.path.getsize(file_path)
                    file_extension_type = file_extension[1:].lower() if file_extension else 'unknown'
                    
                    # Prepare metadata
                    file_metadata = {
                        'id': f"{batch_id}_{new_filename}",
                        'batch_id': batch_id,
                        'institution_id': institution_id,
                        'institution_name': institution_name,
                        'original_filename': file.filename,
                        'saved_filename': new_filename,
                        'file_path': file_path,
                        'file_size': file_size,
                        'file_type': file_extension_type,
                        'upload_date': upload_date,
                        'upload_time': upload_time,
                        'uploaded_at': now.isoformat(),
                        'status': 'uploaded',
                        'processed': False
                    }
                    
                    # Save metadata
                    save_metadata(file_metadata)
                    
                    uploaded_files.append({
                        'original_filename': file.filename,
                        'saved_filename': new_filename,
                        'file_path': file_path,
                        'size': file_size,
                        'type': file_extension_type,
                        'status': 'success',
                        'preview_url': f'/api/files/preview/{safe_institution_name}/{new_filename}'
                    })
                    
                    # Mark for processing
                    file_metadata['needs_processing'] = True
                    
                except Exception as e:
                    failed_files.append({
                        'filename': file.filename,
                        'error': str(e)
                    })
            else:
                failed_files.append({
                    'filename': file.filename,
                    'error': 'File type not allowed'
                })

    # Extract data for preview (don't save to Excel yet)
    extracted_data = []
    if uploaded_files:
        try:
            from ocr_pipeline import process_certificate_file, normalize_extracted_data
            for file_info in uploaded_files:
                file_path = file_info['file_path']
                print(f"Extracting data from: {file_path}")
                data = process_certificate_file(file_path)
                
                # Normalize data to ensure consistent field names
                data = normalize_extracted_data(data)
                
                # Add preview metadata
                data['preview_url'] = file_info['preview_url']
                data['original_filename'] = file_info['original_filename']
                extracted_data.append(data)
        except Exception as e:
            print(f"Extraction error: {str(e)}")
            return jsonify({
                'success': False,
                'error': f'Data extraction failed: {str(e)}'
            }), 500
    
    return jsonify({
        'success': True,
        'batch_id': batch_id,
        'uploaded': len(uploaded_files),
        'failed': len(failed_files),
        'files': uploaded_files,
        'failed_files': failed_files,
        'extracted_data': extracted_data,
        'institution_name': institution_name,
        'message': f'Successfully uploaded {len(uploaded_files)} file(s). Please review and confirm.'
    }), 200


@app.route("/api/institution/uploads", methods=["GET"])
def get_uploads():
    """Get all uploaded files metadata"""
    institution_id = request.args.get('institution_id')
    
    metadata = get_all_metadata()
    
    # Filter by institution if provided
    if institution_id:
        metadata = [m for m in metadata if m.get('institution_id') == institution_id]
    
    # Group by batch
    batches = {}
    for item in metadata:
        batch_id = item.get('batch_id')
        if batch_id not in batches:
            batches[batch_id] = {
                'batch_id': batch_id,
                'institution_name': item.get('institution_name'),
                'uploaded_at': item.get('uploaded_at'),
                'files': [],
                'total_files': 0,
                'processed_files': 0
            }
        batches[batch_id]['files'].append(item)
        batches[batch_id]['total_files'] += 1
        if item.get('processed'):
            batches[batch_id]['processed_files'] += 1
    
    return jsonify({
        'success': True,
        'batches': list(batches.values()),
        'total_batches': len(batches)
    }), 200


@app.route("/api/files/preview/<institution_name>/<filename>", methods=["GET"])
def preview_file(institution_name, filename):
    """Serve uploaded files for preview"""
    try:
        file_path = os.path.join(UPLOAD_FOLDER, 'files', institution_name, filename)
        
        if not os.path.exists(file_path):
            return jsonify({"error": "File not found"}), 404
        
        return send_file(file_path, as_attachment=False)
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/institution/confirm-data", methods=["POST"])
def confirm_and_save_data():
    """Confirm and save extracted data to Excel, and copy original files to verified folder"""
    data = request.get_json() or {}
    
    extracted_data = data.get('extracted_data', [])
    institution_name = data.get('institution_name', 'Unknown Institution')
    batch_id = data.get('batch_id')
    
    if not extracted_data:
        return jsonify({"error": "No data to save"}), 400
    
    try:
        from ocr_pipeline import save_to_excel
        output_dir = str(Path(__file__).parent.parent)
        
        # Save to Excel
        excel_path = save_to_excel(extracted_data, institution_name, output_dir)
        
        # Create verified originals folder
        verified_folder = Path(__file__).parent / 'uploads' / 'verified_originals' / institution_name.replace(' ', '_')
        verified_folder.mkdir(parents=True, exist_ok=True)
        
        # Copy original files to verified folder and store paths in Excel metadata
        copied_files = []
        for item in extracted_data:
            original_filename = item.get('original_filename')
            if original_filename:
                # Find the original file
                safe_institution_name = re.sub(r'[^\w\s-]', '', institution_name)
                safe_institution_name = re.sub(r'[-\s]+', '_', safe_institution_name)
                source_folder = Path(__file__).parent / 'uploads' / 'files' / safe_institution_name
                
                # Look for the file (may have been renamed with batch_id prefix)
                source_file = None
                for file in source_folder.glob('*'):
                    if original_filename in file.name or file.name.endswith(original_filename):
                        source_file = file
                        break
                
                if source_file and source_file.exists():
                    # Copy to verified folder with student ID in filename
                    student_id = item.get('student_id', 'unknown')
                    dest_filename = f"{student_id}_{original_filename}"
                    dest_path = verified_folder / dest_filename
                    
                    import shutil
                    shutil.copy2(source_file, dest_path)
                    
                    copied_files.append({
                        'student_id': student_id,
                        'original_file': original_filename,
                        'verified_path': str(dest_path)
                    })
                    
                    print(f"[SAVE] Copied original file: {original_filename} -> {dest_path}")
        
        if excel_path:
            return jsonify({
                'success': True,
                'excel_file': excel_path,
                'total_records': len(extracted_data),
                'verified_files': copied_files,
                'message': f'Successfully saved {len(extracted_data)} record(s) to Excel and {len(copied_files)} original file(s)'
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to save data to Excel'
            }), 500
            
    except Exception as e:
        print(f"Save error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Failed to save data: {str(e)}'
        }), 500


@app.route("/api/auth/institution/signup", methods=["POST"])
def institution_signup():
    """Handle institution signup - creates account in Supabase and generates SHA256 hash"""
    if not supabase:
        return jsonify({"error": "Database not configured"}), 503

    data = request.get_json() or {}
    
    # Validate required fields
    required_fields = ['institutionName', 'institutionCode', 'email', 'password', 'website']
    missing_fields = [field for field in required_fields if not data.get(field)]
    
    if missing_fields:
        return jsonify({"error": f"Missing required fields: {', '.join(missing_fields)}"}), 400
    
    try:
        # Check if institution already exists
        existing = supabase.table('institutions').select('*').eq('email', data['email']).execute()
        if existing.data and len(existing.data) > 0:
            return jsonify({"error": "Institution with this email already exists"}), 409
        
        existing_code = supabase.table('institutions').select('*').eq('institution_code', data['institutionCode']).execute()
        if existing_code.data and len(existing_code.data) > 0:
            return jsonify({"error": "Institution code already exists"}), 409
        
        # Generate SHA256 hash of institution data (for verification)
        hash_data = {
            'institutionName': data['institutionName'],
            'institutionCode': data['institutionCode'],
            'email': data['email'],
            'website': data['website'],
            'establishedYear': data.get('establishedYear', ''),
            'timestamp': datetime.utcnow().isoformat()
        }
        hash_string = json.dumps(hash_data, sort_keys=True)
        institution_hash = hashlib.sha256(hash_string.encode()).hexdigest()
        
        # Insert institution record
        institution_data = {
            'institution_name': data['institutionName'],
            'institution_code': data['institutionCode'],
            'email': data['email'],
            'password': data['password'],  # In production, hash this!
            'website': data['website'],
            'address': data.get('address', ''),
            'city': data.get('city', ''),
            'state': data.get('state', ''),
            'pincode': data.get('pincode', ''),
            'established_year': data.get('establishedYear', ''),
            'institution_type': data.get('institutionType', ''),
            'affiliated_board': data.get('affiliatedBoard', ''),
            'contact_person_name': data.get('contactPersonName', ''),
            'contact_person_email': data.get('contactPersonEmail', ''),
            'contact_person_phone': data.get('contactPersonPhone', ''),
            'contact_person_designation': data.get('contactPersonDesignation', ''),
            'institution_hash': institution_hash,
            'role': 'institution',
            'verified': False,
            'created_at': datetime.utcnow().isoformat()
        }
        
        result = supabase.table('institutions').insert(institution_data).execute()
        if not (result.data and len(result.data) > 0):
            return jsonify({"error": "Failed to create institution account"}), 500
        
        return jsonify({
            "success": True,
            "message": "Institution registered successfully!",
            "institutionHash": institution_hash,
            "institutionCode": data['institutionCode']
        }), 201
            
    except Exception as e:
        print(f"Institution signup error: {str(e)}")
        return jsonify({"error": f"Registration failed: {str(e)}"}), 500


@app.route("/api/auth/institution/signin", methods=["POST"])
def institution_signin():
    """Handle institution signin - verifies email, password, and hash key"""
    if not supabase:
        return jsonify({"error": "Database not configured"}), 503

    data = request.get_json() or {}
    
    # Validate required fields
    required_fields = ['email', 'password', 'hashKey']
    missing_fields = [field for field in required_fields if not data.get(field)]
    
    if missing_fields:
        return jsonify({"error": f"Missing required fields: {', '.join(missing_fields)}"}), 400
    
    try:
        # Find institution by email
        result = supabase.table('institutions').select('*').eq('email', data['email']).execute()
        
        if not result.data or len(result.data) == 0:
            return jsonify({"error": "Invalid email or password"}), 401
        
        institution = result.data[0]
        
        # Verify password (In production, use proper password hashing!)
        if institution.get('password') != data['password']:
            return jsonify({"error": "Invalid email or password"}), 401
        
        # Verify hash key
        if institution.get('institution_hash') != data['hashKey']:
            return jsonify({"error": "Error in hash key, please provide correct hash key"}), 401
        
        # Successful authentication
        return jsonify({
            "success": True,
            "message": "Sign in successful",
            "institution": {
                "id": institution.get('id'),
                "institutionName": institution.get('institution_name'),
                "institutionCode": institution.get('institution_code'),
                "email": institution.get('email'),
                "contactPersonName": institution.get('contact_person_name'),
                "verified": institution.get('verified', False),
                "createdAt": institution.get('created_at')
            }
        }), 200
            
    except Exception as e:
        print(f"Institution signin error: {str(e)}")
        return jsonify({"error": f"Sign in failed: {str(e)}"}), 500


@app.route("/api/verify-upload", methods=["POST"])
def verify_certificate_upload():
    """Verify uploaded certificate: extract OCR data and match with database"""
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "Empty filename"}), 400
    
    if not allowed_file(file.filename):
        return jsonify({"error": "File type not allowed"}), 400

    try:
        # Save uploaded file temporarily
        uploads_dir = Path(__file__).parent / 'uploads' / 'verify'
        uploads_dir.mkdir(parents=True, exist_ok=True)
        
        filename = secure_filename(file.filename)
        unique_name = f"{uuid.uuid4().hex}_{filename}"
        save_path = uploads_dir / unique_name
        file.save(str(save_path))
        
        # Extract data using Gemini AI
        print(f"Extracting data from: {save_path}")
        from ocr_pipeline import normalize_extracted_data
        extracted_data = process_certificate_file(str(save_path))
        
        # Normalize data to match Excel structure (CRITICAL for matching)
        extracted_data = normalize_extracted_data(extracted_data)
        
        print(f"[VERIFICATION] Normalized extracted data fields: {list(extracted_data.keys())}")
        
        # Check if extraction had errors
        if extracted_data.get('error'):
            return jsonify({
                "success": False,
                "error": "Failed to extract certificate data",
                "details": extracted_data.get('error')
            }), 400
        
        # Search Excel files for matching certificate
        verification_result = {
            "extracted_data": extracted_data,
            "database_match": None,
            "match_percentage": 0,
            "verification_status": "unverified",
            "template_match": False,
            "alert_sent": False,
            "matched_file": None
        }
        
        try:
            # ===== EXCEL VERIFICATION MATCHING =====
            # CRITICAL: This section matches extracted data against institution Excel files
            # Field names MUST be consistent with:
            #   1. ocr_pipeline.py - get_standard_field_names()
            #   2. ocr_pipeline.py - save_to_excel() columns
            #   3. calculate_match_percentage_excel() field comparisons
            # Standard fields: student_name, student_id, university_name, degree_type,
            #                 course_name, specialization, cgpa, year_of_passing, etc.
            # ==========================================
            
            # Search for matching student in Excel files
            student_id = extracted_data.get('student_id', '').strip()
            student_name = extracted_data.get('student_name', '').strip()
            
            print(f"[VERIFICATION] Searching for student_id: '{student_id}', name: '{student_name}'")
            
            if student_id or student_name:
                # Get all Excel files in the project root (029 folder)
                project_root = Path(__file__).parent.parent
                excel_files = list(project_root.glob('*.xlsx'))
                
                print(f"[VERIFICATION] Found {len(excel_files)} Excel file(s) to search")
                
                best_match = None
                best_match_score = 0
                matched_file_name = None
                
                for excel_file in excel_files:
                    try:
                        print(f"[VERIFICATION] Searching in file: {excel_file.name}")
                        
                        # Read Excel file
                        import openpyxl
                        wb = openpyxl.load_workbook(excel_file, data_only=True)
                        ws = wb.active
                        
                        # Get header row to map column names
                        headers = {}
                        raw_headers = {}
                        for idx, cell in enumerate(ws[1], start=1):
                            if cell.value:
                                # Normalize header names (handle both "Student Name" and "student_name")
                                normalized = cell.value.lower().replace(' ', '_').replace('-', '_')
                                headers[normalized] = idx
                                raw_headers[idx] = cell.value
                        
                        print(f"[VERIFICATION] Excel headers: {list(headers.keys())}")
                        
                        # Create column name mapping to handle both old and new formats
                        # Map variations to standardized field names
                        column_mapping = {
                            # Old format mappings
                            'unicgpa': 'cgpa',
                            'grade': None,  # Ignore
                            'completion_date': None,  # Ignore
                            # New friendly format mappings
                            'university': 'university_name',
                            'course': 'course_name',
                            'semester': 'semester',
                            'year_of_passing': 'year_of_passing',
                            'issue_date': 'issue_date',
                            'certificate_number': 'certificate_number',
                            'issuing_authority': 'issuing_authority',
                            'subject_wise_grades': 'subject_grades',
                            'subjectwise_grades': 'subject_grades',
                            'source_file': 'file_name',
                            'upload_date': 'upload_date',
                            'extracted_at': 'extracted_at'
                        }
                        
                        # Search through rows
                        for row_idx in range(2, ws.max_row + 1):
                            row_data = {}
                            for col_name, col_idx in headers.items():
                                cell_value = ws.cell(row_idx, col_idx).value
                                
                                # Apply column mapping
                                if col_name in column_mapping:
                                    mapped_name = column_mapping[col_name]
                                    if mapped_name is None:
                                        continue  # Skip this column
                                    col_name = mapped_name
                                
                                row_data[col_name] = str(cell_value) if cell_value else ''
                            
                            # Check if this row matches the student ID
                            excel_student_id = row_data.get('student_id', '').strip()
                            
                            if excel_student_id and excel_student_id == student_id:
                                print(f"[VERIFICATION] Found matching student_id in row {row_idx}")
                                
                                # Calculate match percentage
                                match_score = calculate_match_percentage_excel(extracted_data, row_data)
                                
                                print(f"[VERIFICATION] Match score: {match_score}%")
                                
                                if match_score > best_match_score:
                                    best_match_score = match_score
                                    best_match = row_data
                                    matched_file_name = excel_file.name
                        
                        wb.close()
                        
                    except Exception as file_error:
                        print(f"[VERIFICATION] Error reading {excel_file.name}: {str(file_error)}")
                        continue
                
                if best_match:
                    verification_result["database_match"] = {
                        "student_id": best_match.get('student_id', ''),
                        "student_name": best_match.get('student_name', ''),
                        "university": best_match.get('university_name', ''),
                        "course": best_match.get('course_name', ''),
                        "degree_type": best_match.get('degree_type', ''),
                        "cgpa": best_match.get('cgpa', ''),
                        "year_of_passing": best_match.get('year_of_passing', ''),
                        "issuing_authority": best_match.get('issuing_authority', '')
                    }
                    verification_result["match_percentage"] = best_match_score
                    verification_result["matched_file"] = matched_file_name
                    
                    # Determine verification status based on match percentage
                    if best_match_score >= 85:
                        verification_result["verification_status"] = "verified"
                        print(f"[VERIFICATION] Status: VERIFIED (≥85%)")
                    elif best_match_score >= 70:
                        verification_result["verification_status"] = "semi-verified"
                        verification_result["template_match"] = True
                        verification_result["alert_sent"] = True
                        print(f"[VERIFICATION] Status: SEMI-VERIFIED (70-84%)")
                    else:
                        verification_result["verification_status"] = "mismatch"
                        print(f"[VERIFICATION] Status: MISMATCH (<70%)")
                else:
                    # No matching record found in any Excel file
                    verification_result["verification_status"] = "not_found"
                    print(f"[VERIFICATION] Status: NOT FOUND - No records with student_id '{student_id}' in any Excel file")
            else:
                print(f"[VERIFICATION] ERROR: No student_id or name extracted from certificate")
                verification_result["verification_status"] = "error"
                verification_result["error_message"] = "No student ID or name could be extracted from the certificate"
                    
        except Exception as search_error:
            print(f"[VERIFICATION] Search error: {str(search_error)}")
            verification_result["verification_status"] = "error"
            verification_result["error_message"] = f"Search error: {str(search_error)}"
        
        # Clean up temporary file
        try:
            os.remove(str(save_path))
        except:
            pass
        
        return jsonify({
            "success": True,
            "result": verification_result
        }), 200
        
    except Exception as e:
        print(f"Verification error: {str(e)}")
        return jsonify({
            "success": False,
            "error": "Verification failed",
            "details": str(e)
        }), 500


def calculate_match_percentage(extracted_data, db_student, db_profile):
    """Calculate percentage match between extracted data and database records"""
    total_fields = 0
    matched_fields = 0
    
    # Compare student_id
    total_fields += 1
    if extracted_data.get('student_id', '').strip().lower() == db_student.get('student_id', '').strip().lower():
        matched_fields += 1
    
    # Compare student name
    total_fields += 1
    extracted_name = extracted_data.get('student_name', '').strip().lower()
    db_name = db_profile.get('name', '').strip().lower()
    if extracted_name and db_name:
        # Fuzzy match for names
        if extracted_name == db_name or extracted_name in db_name or db_name in extracted_name:
            matched_fields += 1
    
    # Compare university
    total_fields += 1
    extracted_uni = extracted_data.get('university_name', '').strip().lower()
    db_uni = db_profile.get('university', '').strip().lower()
    if extracted_uni and db_uni:
        if extracted_uni == db_uni or extracted_uni in db_uni or db_uni in extracted_uni:
            matched_fields += 1
    
    # Compare course
    total_fields += 1
    extracted_course = extracted_data.get('course_name', '').strip().lower()
    db_course = db_profile.get('course', '').strip().lower()
    if extracted_course and db_course:
        if extracted_course == db_course or extracted_course in db_course or db_course in extracted_course:
            matched_fields += 1
    
    # Calculate percentage
    if total_fields > 0:
        percentage = (matched_fields / total_fields) * 100
        return round(percentage, 2)
    
    return 0.0


def calculate_match_percentage_excel(extracted_data, excel_row):
    """
    Calculate percentage match between extracted data and Excel row data
    
    CRITICAL: Field names must match exactly with:
    - ocr_pipeline.py get_standard_field_names()
    - Excel column headers defined in save_to_excel()
    
    Standard fields: student_name, student_id, university_name, degree_type,
                    course_name, specialization, cgpa, year_of_passing, 
                    issue_date, certificate_number, issuing_authority
    
    Args:
        extracted_data: Normalized data from verification upload
        excel_row: Row data from Excel file (with matching column names)
        
    Returns:
        float: Match percentage (0-100)
    """
    total_fields = 0
    matched_fields = 0
    match_details = []  # For debugging
    
    # Compare student_id (exact match - highest priority)
    total_fields += 2  # Weight this more heavily
    extracted_id = extracted_data.get('student_id', '').strip().lower()
    excel_id = excel_row.get('student_id', '').strip().lower()
    if extracted_id == excel_id:
        matched_fields += 2
        match_details.append(f"✓ student_id: '{extracted_id}' == '{excel_id}'")
    else:
        match_details.append(f"✗ student_id: '{extracted_id}' != '{excel_id}'")
    
    # Compare student name
    total_fields += 1
    extracted_name = extracted_data.get('student_name', '').strip().lower()
    excel_name = excel_row.get('student_name', '').strip().lower()
    if extracted_name and excel_name:
        # Fuzzy match for names
        if extracted_name == excel_name or extracted_name in excel_name or excel_name in extracted_name:
            matched_fields += 1
            match_details.append(f"✓ student_name: '{extracted_name}' ≈ '{excel_name}'")
        else:
            match_details.append(f"✗ student_name: '{extracted_name}' != '{excel_name}'")
    else:
        match_details.append(f"⊘ student_name: Empty field")
    
    # Compare university
    total_fields += 1
    extracted_uni = extracted_data.get('university_name', '').strip().lower()
    excel_uni = excel_row.get('university_name', '').strip().lower()
    if extracted_uni and excel_uni:
        if extracted_uni == excel_uni or extracted_uni in excel_uni or excel_uni in extracted_uni:
            matched_fields += 1
            match_details.append(f"✓ university_name: '{extracted_uni}' ≈ '{excel_uni}'")
        else:
            match_details.append(f"✗ university_name: '{extracted_uni}' != '{excel_uni}'")
    else:
        match_details.append(f"⊘ university_name: Empty field - extracted:'{extracted_uni}', excel:'{excel_uni}'")
    
    # Compare course
    total_fields += 1
    extracted_course = extracted_data.get('course_name', '').strip().lower()
    excel_course = excel_row.get('course_name', '').strip().lower()
    if extracted_course and excel_course:
        if extracted_course == excel_course or extracted_course in excel_course or excel_course in extracted_course:
            matched_fields += 1
            match_details.append(f"✓ course_name: '{extracted_course}' ≈ '{excel_course}'")
        else:
            match_details.append(f"✗ course_name: '{extracted_course}' != '{excel_course}'")
    else:
        match_details.append(f"⊘ course_name: Empty field")
    
    # Compare degree type
    total_fields += 1
    extracted_degree = extracted_data.get('degree_type', '').strip().lower()
    excel_degree = excel_row.get('degree_type', '').strip().lower()
    if extracted_degree and excel_degree:
        if extracted_degree == excel_degree or extracted_degree in excel_degree or excel_degree in extracted_degree:
            matched_fields += 1
            match_details.append(f"✓ degree_type: '{extracted_degree}' ≈ '{excel_degree}'")
        else:
            match_details.append(f"✗ degree_type: '{extracted_degree}' != '{excel_degree}'")
    else:
        match_details.append(f"⊘ degree_type: Empty field")
    
    # Compare CGPA (allow small variance)
    total_fields += 1
    extracted_cgpa = extracted_data.get('cgpa', '').strip()
    excel_cgpa = excel_row.get('cgpa', '').strip()
    if extracted_cgpa and excel_cgpa:
        try:
            extracted_cgpa_float = float(extracted_cgpa)
            excel_cgpa_float = float(excel_cgpa)
            # Consider match if within 0.1 difference
            if abs(extracted_cgpa_float - excel_cgpa_float) <= 0.1:
                matched_fields += 1
                match_details.append(f"✓ cgpa: {extracted_cgpa} ≈ {excel_cgpa}")
            else:
                match_details.append(f"✗ cgpa: {extracted_cgpa} != {excel_cgpa}")
        except:
            if extracted_cgpa == excel_cgpa:
                matched_fields += 1
                match_details.append(f"✓ cgpa: '{extracted_cgpa}' == '{excel_cgpa}'")
            else:
                match_details.append(f"✗ cgpa: '{extracted_cgpa}' != '{excel_cgpa}'")
    else:
        match_details.append(f"⊘ cgpa: Empty field")
    
    # Compare year of passing
    total_fields += 1
    extracted_year = extracted_data.get('year_of_passing', '').strip()
    excel_year = excel_row.get('year_of_passing', '').strip()
    if extracted_year and excel_year:
        if extracted_year == excel_year:
            matched_fields += 1
            match_details.append(f"✓ year_of_passing: '{extracted_year}' == '{excel_year}'")
        else:
            match_details.append(f"✗ year_of_passing: '{extracted_year}' != '{excel_year}'")
    else:
        match_details.append(f"⊘ year_of_passing: Empty field")
    
    # Compare issuing authority
    total_fields += 1
    extracted_issuer = extracted_data.get('issuing_authority', '').strip().lower()
    excel_issuer = excel_row.get('issuing_authority', '').strip().lower()
    if extracted_issuer and excel_issuer:
        if extracted_issuer == excel_issuer or extracted_issuer in excel_issuer or excel_issuer in extracted_issuer:
            matched_fields += 1
            match_details.append(f"✓ issuing_authority: '{extracted_issuer}' ≈ '{excel_issuer}'")
        else:
            match_details.append(f"✗ issuing_authority: '{extracted_issuer}' != '{excel_issuer}'")
    else:
        match_details.append(f"⊘ issuing_authority: Empty field")
    
    # Calculate percentage
    percentage = 0.0
    if total_fields > 0:
        percentage = (matched_fields / total_fields) * 100
        percentage = round(percentage, 2)
    
    # Print detailed match breakdown
    print(f"[MATCH DETAILS] Field-by-field comparison:")
    for detail in match_details:
        print(f"  {detail}")
    print(f"[MATCH SUMMARY] Matched {matched_fields}/{total_fields} fields = {percentage}%")
    
    return percentage


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)


# Endpoint: certificate verification using Gemini AI OCR
@app.route("/api/verify/certificate", methods=["POST"])
def verify_certificate_route():
    """Verify certificate using Gemini AI and save data to JSON/Excel"""
    # Expecting multipart/form-data with file field 'file'
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "Empty filename"}), 400

    uploads_dir = Path(__file__).parent / 'uploads'
    uploads_dir.mkdir(parents=True, exist_ok=True)

    filename = secure_filename(file.filename)
    # ensure unique
    unique_name = f"{uuid.uuid4().hex}_{filename}"
    save_path = uploads_dir / unique_name
    file.save(str(save_path))

    try:
        # Output directory: project root (029 folder)
        output_dir = str(Path(__file__).parent.parent)
        
        # Process single file using Gemini AI
        certificate_data = process_certificate_file(str(save_path))
        
        # Get institution name from form data or use 'Unknown'
        institution_name = request.form.get('institution_name', 'Unknown_Institution')
        
        # Save to Excel
        from ocr_pipeline import save_to_excel
        excel_path = save_to_excel([certificate_data], institution_name, output_dir)
        
        result = {
            'extracted_data': certificate_data,
            'excel_file': excel_path,
            'verdict': 'extracted',
            'confidence': 1.0 if not certificate_data.get('error') else 0.0
        }
        
        # Clean up uploaded file after processing
        # os.remove(str(save_path))  # Uncomment to delete after processing
        
        return jsonify({
            "success": True,
            "result": result,
            "message": "Certificate data extracted and saved successfully"
        }), 200
        
    except Exception as e:
        print('Verification error:', str(e))
        return jsonify({
            "error": "Verification failed",
            "detail": str(e)
        }), 500