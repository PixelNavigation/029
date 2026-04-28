import os
import re
import json
import mimetypes
from datetime import datetime
from pathlib import Path
from google import genai
from google.genai import types
import pandas as pd
import hashlib
import uuid


def setup_gemini():
    api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    if not api_key:
        raise ValueError("GEMINI_API_KEY not found in environment variables")
    return genai.Client(api_key=api_key)


def extract_certificate_data_with_gemini(file_path):
    """
    Use Gemini AI to extract all certificate details from a file (image, PDF, etc.).
    
    Returns:
        dict: Extracted certificate data
    """
    try:
        client = setup_gemini()
        
        # Determine file type and handle accordingly
        file_ext = os.path.splitext(file_path)[1].lower()
        
        # For PDFs, upload directly to Gemini
        if file_ext == '.pdf':
            uploaded_file = client.files.upload(file=file_path)
            content_input = uploaded_file
        else:
            mime_type, _ = mimetypes.guess_type(file_path)
            if not mime_type:
                mime_type = "image/jpeg"
            with open(file_path, "rb") as f:
                image_bytes = f.read()
            content_input = types.Part.from_bytes(data=image_bytes, mime_type=mime_type)
        
        # Prompt for Gemini to extract certificate details
        prompt = """
        Analyze this certificate/document image and extract ALL relevant information in a structured format.
        
        Please extract the following details if present:
        1. Student/Recipient Name
        2. Student ID or Registration Number
        3. University/Institution Name (e.g., "Osmania University")
        4. Course/Program (just the degree type: "B.E.", "B.Tech", "M.Sc", "M.Tech", "Diploma", etc.)
        5. Specialization/Branch (e.g., "CSE", "ECE", "Computer Science", "Mechanical")
        6. Semester (just the number: "5", "3", "1", etc.)
        7. CGPA/Percentage
        8. Year of Passing (CRITICAL: Extract the year when the student completed/passed the course. Look for phrases like "Year of Passing", "Passed in", "Completion Year", "Year:", or any 4-digit year near completion/graduation context. Common formats: "2023", "2022-23", "Year: 2024". This is MANDATORY - look carefully!)
        9. Date of Issue
        10. Certificate Number
        11. Issuing Authority (the organization name like "Osmania University")
        12. Subject-wise Grades (extract all subjects with their individual grades/marks)
        
        Return the data in JSON format with these keys:
        {
            "student_name": "",
            "student_id": "",
            "university_name": "",
            "course_name": "",
            "specialization": "",
            "semester": "",
            "cgpa": "",
            "year_of_passing": "",
            "issue_date": "",
            "certificate_number": "",
            "issuing_authority": "",
            "subject_grades": [
                {"subject_name": "", "grade": "", "marks": "", "credits": ""}
            ],
            "raw_text": ""
        }
        
        CRITICAL EXTRACTION RULES:
        - course_name: ONLY the degree type without specialization (e.g., "B.E.", "B.Tech", "M.Sc")
        - specialization: ONLY the branch/stream (e.g., "CSE", "Computer Science Engineering", "ECE")
        - semester: ONLY the number (e.g., "5", "3", "1")
        - DO NOT mix course, specialization, and semester - keep them completely separate
        - issuing_authority: Organization name (e.g., "Osmania University")
        - If any field is not found, leave it as an empty string or empty array
        """
        
        # Generate content with Gemini
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=[prompt, content_input],
        )
        
        # Parse the response
        response_text = response.text.strip()
        
        # Extract JSON from response (handle markdown code blocks)
        if "```json" in response_text:
            json_start = response_text.find("```json") + 7
            json_end = response_text.find("```", json_start)
            json_str = response_text[json_start:json_end].strip()
        elif "```" in response_text:
            json_start = response_text.find("```") + 3
            json_end = response_text.find("```", json_start)
            json_str = response_text[json_start:json_end].strip()
        else:
            json_str = response_text
        
        # Parse JSON
        try:
            certificate_data = json.loads(json_str)
        except json.JSONDecodeError:
            # Fallback: create structured data from text
            certificate_data = {
                "student_name": "",
                "student_id": "",
                "university_name": "",
                "course_name": "",
                "specialization": "",
                "semester": "",
                "cgpa": "",
                "year_of_passing": "",
                "issue_date": "",
                "certificate_number": "",
                "issuing_authority": "",
                "subject_grades": [],
                "raw_text": response_text
            }
        
        # Build degree_type from course_name + semester
        course = certificate_data.get('course_name', '').strip()
        semester = certificate_data.get('semester', '').strip()
        if course and semester:
            certificate_data['degree_type'] = f"{course} Semester {semester}"
        elif course:
            certificate_data['degree_type'] = course
        else:
            certificate_data['degree_type'] = ""
        
        # Add metadata
        certificate_data['extracted_at'] = datetime.utcnow().isoformat()
        certificate_data['upload_date'] = datetime.now().strftime('%Y-%m-%d')
        certificate_data['file_path'] = str(file_path)
        certificate_data['file_name'] = os.path.basename(file_path)
        
        return certificate_data
        
    except Exception as e:
        print(f"Gemini extraction error: {str(e)}")
        return {
            "error": str(e),
            "student_name": "",
            "student_id": "",
            "degree_type": "",
            "university_name": "",
            "raw_text": "",
            "subject_grades": [],
            "upload_date": datetime.now().strftime('%Y-%m-%d'),
            "extracted_at": datetime.utcnow().isoformat()
        }
    
def create_certificate_hash(certificate_data, hash_salt=None):
    """Create the canonical SHA-256 hash for a certificate.

    H = SHA256(U || norm(Sn) || norm(Sid) || norm(C) || norm(Scores))

    - U: UUIDv4 cryptographic salt
    - Sn: student name
    - Sid: student ID
    - C: CGPA
    - Scores: subject-specific marks/grades
    """

    def _norm(value: str) -> str:
        if value is None:
            return ""
        collapsed = re.sub(r"\s+", " ", str(value)).strip()
        return collapsed.lower()

    def _normalize_scores(subject_grades):
        if not subject_grades:
            return ""

        if isinstance(subject_grades, str):
            return _norm(subject_grades)

        normalized_items = []
        for item in subject_grades:
            if not isinstance(item, dict):
                continue
            subject_name = _norm(item.get("subject_name", ""))
            score_value = item.get("marks") or item.get("grade") or ""
            score = _norm(score_value)
            if subject_name or score:
                normalized_items.append((subject_name, score))

        normalized_items.sort(key=lambda x: x[0])
        return "|".join([f"{name}:{score}" for name, score in normalized_items])

    salt = hash_salt or certificate_data.get("hash_salt")
    if not salt:
        salt = str(uuid.uuid4())

    student_name = _norm(certificate_data.get("student_name", ""))
    student_id = _norm(certificate_data.get("student_id", ""))
    cgpa_value = certificate_data.get("cgpa") or certificate_data.get("grade") or ""
    cgpa = _norm(cgpa_value)
    scores = _normalize_scores(certificate_data.get("subject_grades", []))

    canonical_data = "|".join([salt, student_name, student_id, cgpa, scores])
    encoded_data = canonical_data.encode("utf-8")
    sha256_hash = hashlib.sha256(encoded_data).hexdigest()

    if isinstance(certificate_data, dict) and "hash_salt" not in certificate_data:
        certificate_data["hash_salt"] = salt

    return "0x" + sha256_hash


def save_to_excel(data_list, institution_name, output_dir='e:\\SIH 2025\\029'):
    """
    Save extracted certificate data to a neat Excel file based on institution name
    
    Args:
        data_list: List of dictionaries containing certificate data
        institution_name: Name of the institution
        output_dir: Directory to save the file (default: project root)
    
    Returns:
        str: Path to saved Excel file
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        # Sanitize institution name for filename
        safe_institution_name = re.sub(r'[^\w\s-]', '', institution_name).strip()
        safe_institution_name = re.sub(r'[-\s]+', '_', safe_institution_name)
        
        # Create filename based on institution name only (one file per college)
        filename = f"{safe_institution_name}_Certificates.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # Define clean column order - MUST match verification field names
        columns = [
            'student_name', 
            'student_id', 
            'university_name', 
            'course_name',
            'specialization',
            'semester',
            'degree_type', 
            'cgpa', 
            'year_of_passing', 
            'issue_date', 
            'certificate_number', 
            'issuing_authority', 
            'subject_grades', 
            'file_name',
            'upload_date', 
            'extracted_at',
            'hash_salt',
            'blockchain_hash'
        ]
        
        # Define user-friendly header names for Excel
        header_names = {
            'student_name': 'Student Name',
            'student_id': 'Student ID',
            'university_name': 'University',
            'course_name': 'Course',
            'specialization': 'Specialization',
            'semester': 'Semester',
            'degree_type': 'Degree Type',
            'cgpa': 'CGPA',
            'year_of_passing': 'Year of Passing',
            'issue_date': 'Issue Date',
            'certificate_number': 'Certificate Number',
            'issuing_authority': 'Issuing Authority',
            'subject_grades': 'Subject-wise Grades',
            'file_name': 'Source File',
            'upload_date': 'Upload Date',
            'extracted_at': 'Extracted At',
            'hash_salt': 'Hash Salt',
            'blockchain_hash': 'Blockchain Hash'
        }
        
        # Prepare data with consistent columns
        prepared_data = []
        for data in data_list:
            row = {}
            for col in columns:
                if col == 'subject_grades':
                    # Format subject grades as readable string
                    subject_grades = data.get('subject_grades', [])
                    if isinstance(subject_grades, list) and subject_grades:
                        grades_str = '; '.join([f"{sg.get('subject_name', '')}: {sg.get('grade', '')} ({sg.get('marks', '')})" 
                                               for sg in subject_grades if sg.get('subject_name')])
                        row[col] = grades_str
                    else:
                        row[col] = ''
                else:
                    row[col] = data.get(col, '')
            prepared_data.append(row)
        
        # Create DataFrame with specified column order
        df = pd.DataFrame(prepared_data, columns=columns)
        
        # Rename columns to user-friendly names
        df.rename(columns=header_names, inplace=True)
        
        # Check if file exists to append data
        if os.path.exists(filepath):
            existing_df = pd.read_excel(filepath, engine='openpyxl')
            df = pd.concat([existing_df, df], ignore_index=True)
        
        # Save to Excel
        df.to_excel(filepath, index=False, engine='openpyxl')
        
        # Apply formatting
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Set column widths for better readability
        column_widths = {
            'A': 25,  # Student Name
            'B': 18,  # Student ID
            'C': 30,  # University
            'D': 15,  # Course
            'E': 25,  # Specialization
            'F': 10,  # Semester
            'G': 25,  # Degree Type
            'H': 10,  # CGPA
            'I': 15,  # Year of Passing
            'J': 15,  # Issue Date
            'K': 20,  # Certificate Number
            'L': 30,  # Issuing Authority
            'M': 60,  # Subject-wise Grades
            'N': 30,  # Source File
            'O': 15,  # Upload Date
            'P': 20,  # Extracted At
            'Q': 20,  # Hash Salt
            'R': 70   # Blockchain Hash
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Header formatting
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Set header row height
        ws.row_dimensions[1].height = 35
        
        # Format data rows with alternating colors
        light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        data_font = Font(size=11)
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            # Alternating row colors
            fill = light_fill if row_idx % 2 == 0 else PatternFill()
            
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.border = border
                cell.font = data_font
                cell.fill = fill
        
        # Freeze the header row
        ws.freeze_panes = 'A2'
        
        # Add auto-filter to headers
        ws.auto_filter.ref = ws.dimensions
        
        # Save formatted workbook
        wb.save(filepath)
        
        print(f"Data saved to Excel: {filepath}")
        return filepath
        
    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")
        return None


def process_certificate_file(file_path):
    """
    Extract certificate data from a file using Gemini AI
    
    Args:
        file_path: Path to certificate file
    
    Returns:
        dict: Extracted certificate data
    """
    try:
        certificate_data = extract_certificate_data_with_gemini(file_path)
        return certificate_data
    except Exception as e:
        print(f"Certificate processing error: {str(e)}")
        return {
            'error': str(e),
            'student_name': '',
            'student_id': '',
            'university_name': '',
            'file_name': os.path.basename(file_path)
        }


def batch_process_and_save(file_paths, institution_name, output_dir='e:\\SIH 2025\\029'):
    """
    Process multiple certificate files and save to one Excel per institution
    
    Args:
        file_paths: List of file paths to process
        institution_name: Name of the institution
        output_dir: Directory to save output files
    
    Returns:
        dict: Processing results
    """
    try:
        all_data = []
        processed_count = 0
        failed_count = 0
        
        for file_path in file_paths:
            print(f"Processing: {file_path}")
            data = process_certificate_file(file_path)
            
            if data.get('error'):
                failed_count += 1
            else:
                processed_count += 1
            
            all_data.append(data)
        
        # Save all data to Excel
        excel_path = save_to_excel(all_data, institution_name, output_dir)
        
        return {
            'success': True,
            'processed': processed_count,
            'failed': failed_count,
            'total': len(file_paths),
            'excel_file': excel_path,
            'data': all_data
        }
        
    except Exception as e:
        print(f"Batch processing error: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'processed': 0,
            'failed': len(file_paths)
        }


def get_standard_field_names():
    """
    Returns the standard field names used across the system
    This ensures consistency between Excel storage and verification matching
    
    Returns:
        list: Standard field names in order
    """
    return [
        'student_name', 
        'student_id', 
        'university_name', 
        'course_name',
        'specialization',
        'semester',
        'degree_type', 
        'cgpa', 
        'year_of_passing', 
        'issue_date', 
        'certificate_number', 
        'issuing_authority', 
        'subject_grades'
    ]


def normalize_extracted_data(data):
    """
    Normalize extracted data to ensure all standard fields are present
    This ensures compatibility with Excel structure and verification
    
    Args:
        data: Dictionary of extracted certificate data
        
    Returns:
        dict: Normalized data with all standard fields
    """
    standard_fields = get_standard_field_names()
    normalized = {}
    
    for field in standard_fields:
        normalized[field] = data.get(field, '')
    
    # Ensure degree_type is built from course + semester if not present
    if not normalized['degree_type'] and normalized['course_name'] and normalized['semester']:
        normalized['degree_type'] = f"{normalized['course_name']} Semester {normalized['semester']}"
    elif not normalized['degree_type'] and normalized['course_name']:
        normalized['degree_type'] = normalized['course_name']
    
    # Add metadata fields
    normalized['upload_date'] = data.get('upload_date', datetime.now().strftime('%Y-%m-%d'))
    normalized['extracted_at'] = data.get('extracted_at', datetime.utcnow().isoformat())
    normalized['file_name'] = data.get('file_name', '')
    
    # Keep original_filename if present (for institution uploads)
    if 'original_filename' in data:
        normalized['original_filename'] = data['original_filename']
    
    return normalized
