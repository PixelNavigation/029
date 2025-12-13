import os
import uuid
import traceback
from pathlib import Path

from flask import jsonify
from werkzeug.utils import secure_filename

from ocr_pipeline import process_certificate_file, normalize_extracted_data


def verify_certificate_upload(file, calculate_match_percentage_excel):
    """Core logic for certificate verification.

    This function assumes the incoming file has already been validated
    (presence, filename, and extension) by the Flask route. It saves the
    file temporarily, runs OCR + normalization, then searches Excel exports
    in the project root to find the best matching record.
    """
    try:
        # Save uploaded file temporarily
        uploads_dir = Path(__file__).parent / "uploads" / "verify"
        uploads_dir.mkdir(parents=True, exist_ok=True)

        filename = secure_filename(file.filename)
        unique_name = f"{uuid.uuid4().hex}_{filename}"
        save_path = uploads_dir / unique_name
        file.save(str(save_path))

        # Extract data using OCR pipeline
        print(f"[VERIFY] Extracting data from: {save_path}")
        extracted_data = process_certificate_file(str(save_path))
        extracted_data = normalize_extracted_data(extracted_data)

        print(
            f"[VERIFICATION] Normalized extracted data fields: "
            f"{list(extracted_data.keys())}"
        )

        # Check if extraction had errors
        if extracted_data.get("error"):
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "Failed to extract certificate data",
                        "details": extracted_data.get("error"),
                    }
                ),
                400,
            )

        # Initialize verification result structure
        verification_result = {
            "extracted_data": extracted_data,
            "database_match": None,
            "match_percentage": 0,
            "verification_status": "unverified",
            "template_match": False,
            "alert_sent": False,
            "matched_file": None,
        }

        try:
            # Search for matching student in Excel files
            student_id = extracted_data.get("student_id", "").strip()
            student_name = extracted_data.get("student_name", "").strip()

            print(
                f"[VERIFICATION] Searching for student_id: "
                f"'{student_id}', name: '{student_name}'"
            )

            if student_id or student_name:
                project_root = Path(__file__).parent.parent
                excel_files = list(project_root.glob("*.xlsx"))

                print(
                    f"[VERIFICATION] Found {len(excel_files)} "
                    f"Excel file(s) to search"
                )

                best_match = None
                best_match_score = 0
                matched_file_name = None

                import openpyxl

                for excel_file in excel_files:
                    try:
                        print(
                            f"[VERIFICATION] Searching in file: "
                            f"{excel_file.name}"
                        )

                        wb = openpyxl.load_workbook(excel_file, data_only=True)
                        ws = wb.active

                        # Header mapping
                        headers = {}
                        for idx, cell in enumerate(ws[1], start=1):
                            if cell.value:
                                normalized = (
                                    str(cell.value)
                                    .lower()
                                    .replace(" ", "_")
                                    .replace("-", "_")
                                )
                                headers[normalized] = idx

                        # Map variations to standardized field names
                        column_mapping = {
                            "unicgpa": "cgpa",
                            "grade": None,
                            "completion_date": None,
                            "university": "university_name",
                            "course": "course_name",
                            "semester": "semester",
                            "year_of_passing": "year_of_passing",
                            "issue_date": "issue_date",
                            "certificate_number": "certificate_number",
                            "issuing_authority": "issuing_authority",
                            "subject_wise_grades": "subject_grades",
                            "subjectwise_grades": "subject_grades",
                            "source_file": "file_name",
                            "upload_date": "upload_date",
                            "extracted_at": "extracted_at",
                        }

                        # Scan rows
                        for row_idx in range(2, ws.max_row + 1):
                            row_data = {}
                            for col_name, col_idx in headers.items():
                                cell_value = ws.cell(row_idx, col_idx).value

                                if col_name in column_mapping:
                                    mapped_name = column_mapping[col_name]
                                    if mapped_name is None:
                                        continue
                                    col_key = mapped_name
                                else:
                                    col_key = col_name

                                row_data[col_key] = (
                                    str(cell_value) if cell_value is not None else ""
                                )

                            excel_student_id = row_data.get("student_id", "").strip()
                            if excel_student_id and excel_student_id == student_id:
                                match_score, subject_match_info = calculate_match_percentage_excel(
                                    extracted_data, row_data
                                )
                                print(
                                    f"[VERIFICATION] Match score: "
                                    f"{match_score}%"
                                )

                                if match_score > best_match_score:
                                    best_match_score = match_score
                                    best_match = row_data
                                    best_match["subject_match_info"] = subject_match_info
                                    matched_file_name = excel_file.name

                        wb.close()

                    except Exception as file_error:
                        print(
                            f"[VERIFICATION] Error reading {excel_file.name}: "
                            f"{str(file_error)}"
                        )
                        continue

                if best_match:
                    verification_result["database_match"] = {
                        "student_id": best_match.get("student_id", ""),
                        "student_name": best_match.get("student_name", ""),
                        "university": best_match.get("university_name", ""),
                        "course": best_match.get("course_name", ""),
                        "degree_type": best_match.get("degree_type", ""),
                        "cgpa": best_match.get("cgpa", ""),
                        "year_of_passing": best_match.get("year_of_passing", ""),
                        "issuing_authority": best_match.get("issuing_authority", ""),
                    }
                    verification_result["match_percentage"] = best_match_score
                    verification_result["matched_file"] = matched_file_name
                    verification_result["subject_match_info"] = best_match.get("subject_match_info", {})

                    if best_match_score >= 85:
                        verification_result["verification_status"] = "verified"
                    elif best_match_score >= 70:
                        verification_result["verification_status"] = "semi-verified"
                        verification_result["template_match"] = True
                        verification_result["alert_sent"] = True
                    else:
                        verification_result["verification_status"] = "mismatch"
                else:
                    verification_result["verification_status"] = "not_found"
            else:
                print(
                    "[VERIFICATION] ERROR: No student_id or name extracted "
                    "from certificate"
                )
                verification_result["verification_status"] = "error"
                verification_result[
                    "error_message"
                ] = "No student ID or name could be extracted from the certificate"

        except Exception as search_error:
            print(f"[VERIFICATION] Search error: {str(search_error)}")
            verification_result["verification_status"] = "error"
            verification_result["error_message"] = f"Search error: {str(search_error)}"

        # Clean up temporary file
        try:
            os.remove(str(save_path))
        except Exception:
            pass

        return jsonify({"success": True, "result": verification_result}), 200

    except Exception as e:
        print(f"Verification error: {str(e)}")
        traceback.print_exc()
        return (
            jsonify(
                {
                    "success": False,
                    "error": "Verification failed",
                    "details": str(e),
                }
            ),
            500,
        )


def calculate_match_percentage_excel(extracted_data, excel_row):
    """Calculate match percentage between extracted data and Excel row.

    This logic is shared by the verification route in ``app.py`` and lives
    here alongside the core verification handler.
    """
    total_fields = 0
    matched_fields = 0
    match_details = []

    # Compare student_id (exact match - highest priority)
    total_fields += 2  # Weight this more heavily
    extracted_id = extracted_data.get("student_id", "").strip().lower()
    excel_id = excel_row.get("student_id", "").strip().lower()
    if extracted_id == excel_id:
        matched_fields += 2
        match_details.append(f"✓ student_id: '{extracted_id}' == '{excel_id}'")
    else:
        match_details.append(f"✗ student_id: '{extracted_id}' != '{excel_id}'")

    # Compare student name
    total_fields += 1
    extracted_name = extracted_data.get("student_name", "").strip().lower()
    excel_name = excel_row.get("student_name", "").strip().lower()
    if extracted_name and excel_name:
        if (
            extracted_name == excel_name
            or extracted_name in excel_name
            or excel_name in extracted_name
        ):
            matched_fields += 1
            match_details.append(
                f"✓ student_name: '{extracted_name}' ≈ '{excel_name}'"
            )
        else:
            match_details.append(
                f"✗ student_name: '{extracted_name}' != '{excel_name}'"
            )
    else:
        match_details.append("⊘ student_name: Empty field")

    # Compare university
    total_fields += 1
    extracted_uni = extracted_data.get("university_name", "").strip().lower()
    excel_uni = excel_row.get("university_name", "").strip().lower()
    if extracted_uni and excel_uni:
        if (
            extracted_uni == excel_uni
            or extracted_uni in excel_uni
            or excel_uni in extracted_uni
        ):
            matched_fields += 1
            match_details.append(
                f"✓ university_name: '{extracted_uni}' ≈ '{excel_uni}'"
            )
        else:
            match_details.append(
                f"✗ university_name: '{extracted_uni}' != '{excel_uni}'"
            )
    else:
        match_details.append(
            f"⊘ university_name: Empty field - extracted:'{extracted_uni}', excel:'{excel_uni}'"
        )

    # Compare course
    total_fields += 1
    extracted_course = extracted_data.get("course_name", "").strip().lower()
    excel_course = excel_row.get("course_name", "").strip().lower()
    if extracted_course and excel_course:
        if (
            extracted_course == excel_course
            or extracted_course in excel_course
            or excel_course in extracted_course
        ):
            matched_fields += 1
            match_details.append(
                f"✓ course_name: '{extracted_course}' ≈ '{excel_course}'"
            )
        else:
            match_details.append(
                f"✗ course_name: '{extracted_course}' != '{excel_course}'"
            )
    else:
        match_details.append("⊘ course_name: Empty field")

    # Compare degree type
    total_fields += 1
    extracted_degree = extracted_data.get("degree_type", "").strip().lower()
    excel_degree = excel_row.get("degree_type", "").strip().lower()
    if extracted_degree and excel_degree:
        if (
            extracted_degree == excel_degree
            or extracted_degree in excel_degree
            or excel_degree in extracted_degree
        ):
            matched_fields += 1
            match_details.append(
                f"✓ degree_type: '{extracted_degree}' ≈ '{excel_degree}'"
            )
        else:
            match_details.append(
                f"✗ degree_type: '{extracted_degree}' != '{excel_degree}'"
            )
    else:
        match_details.append("⊘ degree_type: Empty field")

    # Compare CGPA (allow small variance)
    total_fields += 1
    extracted_cgpa = extracted_data.get("cgpa", "").strip()
    excel_cgpa = excel_row.get("cgpa", "").strip()
    if extracted_cgpa and excel_cgpa:
        try:
            extracted_cgpa_float = float(extracted_cgpa)
            excel_cgpa_float = float(excel_cgpa)
            if abs(extracted_cgpa_float - excel_cgpa_float) <= 0.1:
                matched_fields += 1
                match_details.append(
                    f"✓ cgpa: {extracted_cgpa} ≈ {excel_cgpa}"
                )
            else:
                match_details.append(
                    f"✗ cgpa: {extracted_cgpa} != {excel_cgpa}"
                )
        except Exception:
            if extracted_cgpa == excel_cgpa:
                matched_fields += 1
                match_details.append(
                    f"✓ cgpa: '{extracted_cgpa}' == '{excel_cgpa}'"
                )
            else:
                match_details.append(
                    f"✗ cgpa: '{extracted_cgpa}' != '{excel_cgpa}'"
                )
    else:
        match_details.append("⊘ cgpa: Empty field")

    # Compare year of passing
    total_fields += 1
    extracted_year = extracted_data.get("year_of_passing", "").strip()
    excel_year = excel_row.get("year_of_passing", "").strip()
    if extracted_year and excel_year:
        if extracted_year == excel_year:
            matched_fields += 1
            match_details.append(
                f"✓ year_of_passing: '{extracted_year}' == '{excel_year}'"
            )
        else:
            match_details.append(
                f"✗ year_of_passing: '{extracted_year}' != '{excel_year}'"
            )
    else:
        match_details.append("⊘ year_of_passing: Empty field")

    # Compare issuing authority
    total_fields += 1
    extracted_issuer = extracted_data.get("issuing_authority", "").strip().lower()
    excel_issuer = excel_row.get("issuing_authority", "").strip().lower()
    if extracted_issuer and excel_issuer:
        if (
            extracted_issuer == excel_issuer
            or extracted_issuer in excel_issuer
            or excel_issuer in extracted_issuer
        ):
            matched_fields += 1
            match_details.append(
                f"✓ issuing_authority: '{extracted_issuer}' ≈ '{excel_issuer}'"
            )
        else:
            match_details.append(
                f"✗ issuing_authority: '{extracted_issuer}' != '{excel_issuer}'"
            )
    else:
        match_details.append("⊘ issuing_authority: Empty field")

    # ===== SUBJECT GRADES COMPARISON =====
    # Compare subject-wise grades and reduce match % if subjects don't match
    extracted_subjects = extracted_data.get("subject_grades", [])
    excel_subjects_str = excel_row.get("subject_grades", "").strip()
    
    subject_match_info = {
        "matched": [],
        "corrected": [],
        "missing_in_excel": [],
        "missing_in_extracted": []
    }
    
    if excel_subjects_str:
        # Parse Excel subject grades string (format: "Subject: Grade (Marks); ...")
        excel_subjects = []
        if isinstance(excel_subjects_str, str) and excel_subjects_str:
            parts = excel_subjects_str.split(';')
            for part in parts:
                part = part.strip()
                if ':' in part:
                    subj_part, grade_part = part.split(':', 1)
                    subj_name = subj_part.strip()
                    grade = grade_part.strip().split('(')[0].strip()
                    
                    # Try to extract credits from the string if present
                    credits = None
                    if '(' in grade_part and ')' in grade_part:
                        # Sometimes format is "Grade (Credits)" or "Grade (Marks)"
                        # We'll try to detect credits pattern
                        pass
                    
                    excel_subjects.append({
                        "subject_name": subj_name,
                        "grade": grade,
                        "credits": credits
                    })
        
        # Compare subjects if both exist
        if extracted_subjects and excel_subjects:
            total_fields += 2  # Weight subject matching
            subject_matches = 0
            
            # Normalize subject names for comparison
            def normalize_subject(name):
                return name.lower().replace('.', '').replace(',', '').replace('  ', ' ').strip()
            
            # Create lookup for exact and fuzzy matching
            excel_subject_map = {normalize_subject(s["subject_name"]): s for s in excel_subjects}
            extracted_subject_names = [normalize_subject(s.get("subject_name", "")) for s in extracted_subjects]
            
            # Track which Excel subjects have been used to prevent duplicate matching
            used_excel_subjects = set()
            
            # Helper function to check if subject is a lab course
            def is_lab_course(name, credits):
                name_lower = name.lower()
                # Lab courses typically have "lab" or "laboratory" in name, or have 1 credit
                is_lab_by_name = 'lab' in name_lower or 'laboratory' in name_lower or 'practical' in name_lower
                is_lab_by_credits = credits and (str(credits) == '1' or credits == 1)
                return is_lab_by_name or is_lab_by_credits
            
            for ext_subj in extracted_subjects:
                ext_name = ext_subj.get("subject_name", "").strip()
                ext_name_norm = normalize_subject(ext_name)
                ext_grade = ext_subj.get("grade", "").strip()
                ext_credits = ext_subj.get("credits", "")
                
                # Convert credits to comparable format
                try:
                    ext_credits_int = int(ext_credits) if ext_credits else None
                except:
                    ext_credits_int = None
                
                # Try exact match first (with same name)
                if ext_name_norm in excel_subject_map and ext_name_norm not in used_excel_subjects:
                    excel_match = excel_subject_map[ext_name_norm]
                    subject_matches += 1
                    used_excel_subjects.add(ext_name_norm)
                    subject_match_info["matched"].append({
                        "extracted": ext_name,
                        "excel": excel_match["subject_name"],
                        "extracted_grade": ext_grade,
                        "excel_grade": excel_match["grade"]
                    })
                else:
                    # Try fuzzy match (partial string match) with credit consideration
                    found_match = False
                    best_match = None
                    best_match_score = 0
                    
                    for excel_name_norm, excel_subj in excel_subject_map.items():
                        # Skip if already used
                        if excel_name_norm in used_excel_subjects:
                            continue
                        
                        # Check name similarity
                        is_substring = ext_name_norm in excel_name_norm or excel_name_norm in ext_name_norm
                        similarity_score = similarity_ratio(ext_name_norm, excel_name_norm)
                        
                        if is_substring or similarity_score > 0.7:
                            # Calculate match score considering both name and credits
                            match_score = similarity_score if not is_substring else 0.9
                            
                            # Boost score if credits match or both are lab/theory courses
                            ext_is_lab = is_lab_course(ext_name, ext_credits_int)
                            excel_is_lab = is_lab_course(excel_subj["subject_name"], None)
                            
                            # If both are lab or both are theory, boost the score
                            if ext_is_lab == excel_is_lab:
                                match_score += 0.2
                            
                            # If this is a better match than previous candidates, save it
                            if match_score > best_match_score:
                                best_match_score = match_score
                                best_match = (excel_name_norm, excel_subj)
                    
                    # If we found a good match, use it
                    if best_match and best_match_score > 0.7:
                        excel_name_norm, excel_subj = best_match
                        subject_matches += 0.7  # Partial credit for fuzzy match
                        used_excel_subjects.add(excel_name_norm)
                        subject_match_info["corrected"].append({
                            "extracted": ext_name,
                            "excel": excel_subj["subject_name"],
                            "extracted_grade": ext_grade,
                            "excel_grade": excel_subj["grade"],
                            "similarity": "fuzzy"
                        })
                        found_match = True
                    
                    if not found_match:
                        subject_match_info["missing_in_excel"].append(ext_name)
            
            # Check for subjects in Excel but not in extracted (only unused subjects)
            for excel_subj in excel_subjects:
                excel_name_norm = normalize_subject(excel_subj["subject_name"])
                # Only add to missing if it wasn't used in any match
                if excel_name_norm not in used_excel_subjects:
                    subject_match_info["missing_in_extracted"].append(excel_subj["subject_name"])
            
            # Calculate subject match contribution
            # Compare against the larger of the two lists for fair scoring
            max_subjects = max(len(extracted_subjects), len(excel_subjects))
            if max_subjects > 0:
                subject_match_ratio = subject_matches / max_subjects
                matched_fields += 2 * subject_match_ratio
                match_details.append(
                    f"{'✓' if subject_match_ratio > 0.7 else '⚠' if subject_match_ratio > 0.3 else '✗'} "
                    f"subjects: {subject_matches}/{max_subjects} matched (extracted: {len(extracted_subjects)}, excel: {len(excel_subjects)})"
                )
            else:
                match_details.append("⊘ subjects: No subjects extracted")
        else:
            match_details.append("⊘ subjects: Empty in one or both records")
    else:
        match_details.append("⊘ subjects: No subjects in Excel")

    percentage = 0.0
    if total_fields > 0:
        percentage = (matched_fields / total_fields) * 100
        percentage = round(percentage, 2)

    print("[MATCH DETAILS] Field-by-field comparison:")
    for detail in match_details:
        print(f"  {detail}")
    print(
        f"[MATCH SUMMARY] Matched {matched_fields}/{total_fields} "
        f"fields = {percentage}%"
    )
    
    # Return both percentage and subject match details
    return percentage, subject_match_info


def similarity_ratio(str1, str2):
    """Calculate similarity between two strings using simple character overlap"""
    if not str1 or not str2:
        return 0.0
    set1 = set(str1.lower())
    set2 = set(str2.lower())
    intersection = set1.intersection(set2)
    union = set1.union(set2)
    return len(intersection) / len(union) if len(union) > 0 else 0.0


